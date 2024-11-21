import pandas as pd
from openpyxl import Workbook
from enum import Enum
from typing import Dict, List, Tuple, Optional
import logging
from dataclasses import dataclass
from collections import defaultdict
import math

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants and type definitions
@dataclass
class ProductInfo:
    priority1: str
    volume1: float
    priority2: str 
    volume2: float
    priority3: str
    volume3: float
    priority4: str
    volume4: float
    family: Optional[str] = None
    resin: Optional[str] = None
    shade: Optional[str] = None
    dio: Optional[float] = None
    aging: Optional[float] = None

class Family(Enum):
    GLOSSY = "glossy"
    SEMI_GLOSSY = "semi-glossy"
    MATT = "matt" 
    TEXTURE = "texture"
    STRUCTURE = "structure"

# Family compatibility matrix (simplified to essential checks)
FAMILY_MATRIX = {
    Family.GLOSSY: {
        Family.GLOSSY: 'OK',
        Family.SEMI_GLOSSY: 'OK', 
        Family.MATT: 'TPS',
        Family.TEXTURE: 'OPS',
        Family.STRUCTURE: 'NP'
    },
    Family.SEMI_GLOSSY: {
        Family.GLOSSY: 'OK',
        Family.SEMI_GLOSSY: 'OK',
        Family.MATT: 'TPS',
        Family.TEXTURE: 'OPS',
        Family.STRUCTURE: 'NP'
    },
    Family.MATT: {
        Family.GLOSSY: 'TPS',
        Family.SEMI_GLOSSY: 'TPS',
        Family.MATT: 'OK',
        Family.TEXTURE: 'OK',
        Family.STRUCTURE: 'NP'
    },
    Family.TEXTURE: {
        Family.GLOSSY: 'OPS',
        Family.SEMI_GLOSSY: 'OPS',
        Family.MATT: 'OK',
        Family.TEXTURE: 'OK',
        Family.STRUCTURE: 'OK'
    },
    Family.STRUCTURE: {
        Family.GLOSSY: 'NP',
        Family.SEMI_GLOSSY: 'NP',
        Family.MATT: 'NP',
        Family.TEXTURE: 'OK',
        Family.STRUCTURE: 'OK'
    }
}

# Resin compatibility matrix
RESIN_MATRIX = {
    'A': {'A': True, 'E': True, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True},
    'E': {'A': True, 'E': True, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True},
    'J': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True},
    'M': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True},
    'S': {'A': True, 'E': False, 'J': False, 'M': True, 'S': True, 'Y': True, 'Q': False, 'K': True, 'F': True, 'H': True, 'L': False, 'N': True},
    'Y': {'A': True, 'E': False, 'J': False, 'M': True, 'S': True, 'Y': True, 'Q': False, 'K': True, 'F': True, 'H': True, 'L': False, 'N': True},
    'Q': {'A': True, 'E': False, 'J': False, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': False, 'N': True},
    'K': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': False, 'N': True},
    'F': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': False, 'N': True},
    'H': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True},
    'L': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True},
    'N': {'A': True, 'E': False, 'J': True, 'M': True, 'S': True, 'Y': True, 'Q': True, 'K': True, 'F': True, 'H': True, 'L': True, 'N': True}
}

# Line configurations
LINE_CONFIG = {
    'Line 1': {'shade': 'dark', 'weekly_capacity': 99000},
    'Line 2': {'shade': 'light', 'weekly_capacity': 79200},
    'Line 3': {'shade': 'light', 'weekly_capacity': 13200},
    'Line 4': {'shade': 'dark', 'weekly_capacity': 52800}
}

class ProductionScheduler:
    def __init__(self):
        self.product_cache = {}
        self.line_states = defaultdict(lambda: defaultdict(dict))
        self.errors = defaultdict(list)
        self.schedule = []  # Initialize schedule attribute
        
    def load_prioritized_products(self) -> Dict[str, float]:
        """
        Load products from empty_prioritized_plan_structure.xlsx
        Returns: Dictionary of product codes and their required volumes
        """
        prioritized_products = {}
        try:
            df_prioritized = pd.read_excel('empty_prioritized_plan_structure.xlsx')
            logger.info(f"Available columns in prioritized plan: {df_prioritized.columns.tolist()}")
            
            # Use the exact column names from your file
            product_column = 'mfg pro'
            volume_column = 'production plan'
            
            if product_column not in df_prioritized.columns or volume_column not in df_prioritized.columns:
                logger.error(f"Could not find required columns. Available columns: {df_prioritized.columns.tolist()}")
                return {}
            
            # Create dictionary of products and their volumes
            for _, row in df_prioritized.iterrows():
                if pd.notna(row[volume_column]) and row[volume_column] > 0:
                    product_code = str(row[product_column]).strip()  # Clean any whitespace
                    volume = float(row[volume_column])
                    prioritized_products[product_code] = volume
            
            logger.info(f"Loaded {len(prioritized_products)} prioritized products with their volumes")
            for prod, vol in list(prioritized_products.items())[:5]:  # Log first 5 products for verification
                logger.info(f"Sample product: {prod} with volume: {vol}")
            
            return prioritized_products
        
        except Exception as e:
            logger.error(f"Error loading prioritized products: {str(e)}")
            return {}

    def load_product_data(self) -> None:
        """Load production details for prioritized products from Product_Summary.xlsx"""
        try:
            # First load prioritized products
            prioritized_products = self.load_prioritized_products()
            if not prioritized_products:
                logger.error("No prioritized products found to schedule")
                return
            
            # Then read Product_Summary.xlsx for their production details
            df = pd.read_excel('Product_Summary.xlsx')
            logger.info(f"Available columns in Product_Summary: {df.columns.tolist()}")
            
            for product, required_volume in prioritized_products.items():
                # Find product details in Product_Summary
                product_data = df[df['Product'] == product]
                
                if product_data.empty:
                    logger.warning(f"Product {product} not found in Product_Summary database")
                    continue
                
                # Get the first matching row
                row = product_data.iloc[0]
                priority1 = row['Priority 1'] if pd.notna(row['Priority 1']) else None
                
                if priority1 is None:
                    logger.warning(f"Skipping product {product} due to missing Priority 1")
                    continue
                
                self.product_cache[product] = ProductInfo(
                    priority1=priority1,
                    volume1=required_volume,  # Use volume from prioritized plan
                    priority2=row.get('Priority 2', ''),
                    volume2=0,  # We only use volume1 as per prioritized plan
                    priority3='',
                    volume3=0,
                    priority4='',
                    volume4=0
                )
                
            logger.info(f"Loaded production details for {len(self.product_cache)} products")
                
        except Exception as e:
            logger.error(f"Error loading product data: {str(e)}")
            raise

    def load_family_data(self) -> None:
        """Load and merge family/shade data"""
        try:
            df_family = pd.read_excel('familydatabase.xlsx')
            
            for _, row in df_family.iterrows():
                if row['mfg pro'] in self.product_cache:
                    product = self.product_cache[row['mfg pro']]
                    product.family = Family(row['family'].lower())
                    product.resin = row['resin']
                    product.shade = 'dark' if row['L_value'] < 50 else 'light'
                    
        except Exception as e:
            logger.error(f"Error loading family data: {str(e)}")
            raise

    def validate_line_assignment(self, product: str, line: str) -> bool:
        """
        Validate if product can be assigned to line based on shade and special rules
        """
        # Add validation for nan or empty line values
        if pd.isna(line) or not isinstance(line, str):
            self.errors[product].append(f"Invalid line assignment: {line}")
            return False
        
        # Validate line exists in config
        if line not in LINE_CONFIG:
            self.errors[product].append(f"Unknown line: {line}")
            return False
        
        product_info = self.product_cache[product]
        line_info = LINE_CONFIG[line]
        
        # Strict shade matching
        if product_info.shade != line_info['shade']:
            error_msg = f"Invalid shade assignment: {product_info.shade} product on {line_info['shade']} line"
            self.errors[product].append(error_msg)
            return False
        
        # Modified Line 3 rules
        if line == 'Line 3':
            # Allow if it's the first priority line
            if product_info.priority1 == 'Line 3':
                return True
            
            # For non-priority1 products, allow if volume is less than 12000
            volume = sum([product_info.volume1, product_info.volume2, product_info.volume3, product_info.volume4])
            if volume < 12000:
                return True
            else:
                self.errors[product].append(f"Volume {volume} too large for Line 3")
                return False
        
        return True

    def calculate_changeover_time(self, current_family: Family, next_family: Family) -> float:
        """Calculate changeover time in hours based on family transition type"""
        if not current_family or not next_family:
            return 0
        
        transition_type = FAMILY_MATRIX.get(current_family, {}).get(next_family)
        
        if transition_type == 'TPS':
            return 3.0  # 3 hours for TPS (2 Times PVC + screw c/o)
        elif transition_type == 'OPS':
            return 3.0  # 3 hours for OPS (1 Time PVC + screw c/o)
        elif transition_type == 'OK':
            return 0.5  # 30 minutes for standard changeover
        else:
            return 0.0

    def get_available_capacity(self, week: int, line: str, current_family: Family, next_family: Family) -> float:
        """Calculate available capacity considering changeover time"""
        # Get base weekly capacity in kg
        base_capacity = LINE_CONFIG[line]['weekly_capacity']
        
        # Convert changeover time to kg of lost production
        throughput_per_hour = {
            'Line 1': 600,
            'Line 2': 500,
            'Line 3': 300,
            'Line 4': 500
        }
        
        changeover_time = self.calculate_changeover_time(current_family, next_family)
        lost_production = changeover_time * throughput_per_hour[line]
        
        # Calculate used capacity from existing schedule
        used_capacity = sum(entry[3] for entry in self.schedule 
                           if entry[0] == week and entry[1] == line)
        
        return base_capacity - used_capacity - lost_production

    def check_compatibility(self, week: int, line: str, product: str, current_family: Optional[Family] = None) -> bool:
        """Enhanced compatibility check including changeover considerations"""
        product_info = self.product_cache[product]
        
        # If this is the first product on the line for the week
        if not current_family:
            return True
        
        transition_type = FAMILY_MATRIX.get(current_family, {}).get(product_info.family)
        
        if transition_type == 'NP':
            return False
        
        # Calculate available capacity including changeover time
        available_capacity = self.get_available_capacity(week, line, current_family, product_info.family)
        
        # Check if we have enough capacity for both changeover and production
        if available_capacity <= 0:
            return False
        
        return True

    def get_transition_priority_score(self, current_family: Family, next_family: Family, product_volume: float) -> float:
        """
        Get priority score for family transitions with additional weighting factors
        Returns weighted score considering:
        - Family transition type (OK/OPS/TPS)
        - Product volume
        - Special family handling (Structure/Matt/Texture)
        """
        if not current_family or not next_family:
            base_score = 3  # First product on line
        else:
            transition_type = FAMILY_MATRIX.get(current_family, {}).get(next_family)
            if transition_type == 'OK':
                base_score = 3
            elif transition_type == 'OPS':
                base_score = 2
            elif transition_type == 'TPS':
                base_score = 1
            else:  # 'NP'
                return -1

        # Add priority for special families to ensure they get scheduled
        special_family_bonus = 0
        if next_family in [Family.STRUCTURE, Family.MATT, Family.TEXTURE]:
            special_family_bonus = 0.5  # Bonus for special families
        
        # Volume consideration (normalized to 0-1 range assuming max volume of 30000)
        volume_factor = min(product_volume / 30000, 1.0) * 0.5
        
        return base_score + special_family_bonus + volume_factor

    def optimize_schedule(self, total_weeks: int = 2) -> List:
        """Enhanced scheduling with priority line and family optimization"""
        self.schedule = []
        current_family_on_line = defaultdict(lambda: defaultdict(lambda: None))
        current_resin_on_line = defaultdict(lambda: defaultdict(lambda: None))
        
        # Track scheduled volumes for each product
        scheduled_volumes = defaultdict(float)
        
        # Helper function to check if product is fully scheduled
        def is_fully_scheduled(product: str, info) -> bool:
            total_planned = sum([info.volume1, info.volume2, info.volume3, info.volume4])
            return scheduled_volumes[product] >= total_planned

        # Helper function to add schedule entry
        def add_schedule_entry(week: int, line: str, product: str, volume: float, info, reason: str):
            if volume <= 0:
                return False
            
            schedule_entry = [
                week,
                line,
                product,
                volume,
                info.family,
                info.resin,
                info.shade,
                reason
            ]
            
            self.schedule.append(schedule_entry)
            scheduled_volumes[product] += volume
            current_family_on_line[week][line] = info.family
            current_resin_on_line[week][line] = info.resin
            return True

        # Get all products
        all_products = [(product, info) for product, info in self.product_cache.items()]
        
        # Group products by their priority lines
        line_products = defaultdict(list)
        for product, info in all_products:
            if info.priority1:  # Group by primary priority line
                line_products[info.priority1].append((product, info))
        
        print("\nInitial product distribution:")
        for line, products in line_products.items():
            print(f"{line}: {len(products)} products")
        
        # First pass - schedule priority products
        for week in range(1, total_weeks + 1):
            print(f"\nScheduling Week {week}")
            
            # Track unscheduled products that could go to Line 3
            line3_candidates = []
            
            for line in LINE_CONFIG:
                print(f"\n{line} scheduling:")
                remaining_capacity = LINE_CONFIG[line]['weekly_capacity']
                current_family = current_family_on_line[week][line]
                current_resin = current_resin_on_line[week][line]
                
                # Get products assigned to this line
                priority_products = line_products[line].copy()
                
                # For Line 3, also consider products that couldn't be scheduled on other lines
                if line == 'Line 3':
                    priority_products.extend(line3_candidates)
                
                if not priority_products:
                    print(f"No priority products for {line}")
                    continue
                
                # Sort products by compatibility with current family and resin
                def get_sort_key(product_info):
                    product, info = product_info
                    if not current_family or not current_resin:
                        return (3, info.volume1)  # First product gets highest priority
                    
                    # Check resin compatibility
                    resin_compatible = RESIN_MATRIX.get(current_resin, {}).get(info.resin, False)
                    if not resin_compatible:
                        return (-1, info.volume1)  # Incompatible resin gets lowest priority
                    
                    # Check family compatibility
                    transition_type = FAMILY_MATRIX.get(current_family, {}).get(info.family)
                    if transition_type == 'OK':
                        return (3, info.volume1)
                    elif transition_type == 'OPS':
                        return (2, info.volume1)
                    elif transition_type == 'TPS':
                        return (1, info.volume1)
                    return (0, info.volume1)
                
                # Schedule products for this line
                scheduled_products = []
                priority_products.sort(key=get_sort_key, reverse=True)
                
                for product, info in priority_products:
                    if remaining_capacity <= 0:
                        break
                    
                    # Check resin compatibility
                    if current_resin and not RESIN_MATRIX.get(current_resin, {}).get(info.resin, False):
                        print(f"Skipping {product} - Incompatible resin transition from {current_resin} to {info.resin}")
                        continue
                    
                    if not self.validate_line_assignment(product, line):
                        print(f"Cannot assign {product} to {line} - validation failed")
                        continue
                    
                    # Calculate changeover impact
                    changeover_time = self.calculate_changeover_time(current_family, info.family)
                    throughput = {'Line 1': 600, 'Line 2': 500, 'Line 3': 300, 'Line 4': 500}[line]
                    lost_production = changeover_time * throughput
                    
                    # Check if we can schedule after changeover
                    available_after_changeover = remaining_capacity - lost_production
                    if available_after_changeover <= 0:
                        print(f"Not enough capacity for {product} after changeover")
                        continue
                    
                    # Check if product can be scheduled on Line 3
                    volume = info.volume1
                    if line != 'Line 3' and volume < 12000:  # Changed from 5000 to 12000
                        # Check family and resin compatibility
                        if current_family and current_resin:
                            family_compatible = FAMILY_MATRIX.get(current_family, {}).get(info.family) != 'NP'
                            resin_compatible = RESIN_MATRIX.get(current_resin, {}).get(info.resin, False)
                            if family_compatible and resin_compatible:
                                line3_candidates.append((product, info))
                                continue
                    
                    # Schedule what we can
                    volume_to_schedule = min(info.volume1, available_after_changeover)
                    if volume_to_schedule > 0:
                        add_schedule_entry(week, line, product, volume_to_schedule, info, f"Priority Line: {info.priority1}")
                        scheduled_products.append(product)
                        remaining_capacity -= (volume_to_schedule + lost_production)
                        current_family = info.family
                        current_resin = info.resin
                        current_family_on_line[week][line] = current_family
                        current_resin_on_line[week][line] = current_resin
                        
                        print(f"Scheduled {product} ({info.family.value}): {volume_to_schedule} units")
                        
                        # Remove scheduled volume from product info
                        info.volume1 -= volume_to_schedule
                
                # Remove fully scheduled products from line_products
                line_products[line] = [(p, i) for p, i in line_products[line] 
                                     if p not in scheduled_products and i.volume1 > 0]
        
        # Second pass - fill empty capacity in weeks 1-2
        print("\nFilling empty line capacity in earlier weeks:")
        unscheduled = []
        for line, products in line_products.items():
            for product, info in products:
                if not is_fully_scheduled(product, info):
                    unscheduled.append((product, info))

        if unscheduled:
            # Try to fill empty capacity in weeks 1-2 first
            for week in range(1, total_weeks + 1):
                for line in LINE_CONFIG:
                    remaining_capacity = LINE_CONFIG[line]['weekly_capacity']
                    
                    # Calculate used capacity
                    used_capacity = sum(entry[3] for entry in self.schedule 
                                      if entry[0] == week and entry[1] == line)
                    remaining_capacity -= used_capacity

                    if remaining_capacity > 0:
                        print(f"\nFilling empty capacity in Week {week} - {line} ({remaining_capacity:,.0f} units available)")
                        
                        # Get current family and resin on this line
                        current_family = current_family_on_line[week][line]
                        current_resin = current_resin_on_line[week][line]
                        
                        # Sort unscheduled products by compatibility
                        compatible_products = []
                        for product, info in unscheduled[:]:
                            # Skip if product is already fully scheduled
                            if is_fully_scheduled(product, info):
                                continue
                                
                            if not self.validate_line_assignment(product, line):
                                continue
                                
                            # Check family and resin compatibility
                            family_compatible = True
                            resin_compatible = True
                            
                            if current_family:
                                transition_type = FAMILY_MATRIX.get(current_family, {}).get(info.family)
                                family_compatible = transition_type != 'NP'
                                
                            if current_resin:
                                resin_compatible = RESIN_MATRIX.get(current_resin, {}).get(info.resin, False)
                            
                            if family_compatible and resin_compatible:
                                if line == 'Line 3' and info.priority1 != 'Line 3':
                                    if info.volume1 < 12000:
                                        compatible_products.append((product, info))
                                else:
                                    compatible_products.append((product, info))
                        
                        # Sort compatible products by volume
                        compatible_products.sort(key=lambda x: x[1].volume1, reverse=True)
                        
                        # Try to schedule compatible products
                        for product, info in compatible_products:
                            if remaining_capacity <= 0:
                                break
                                
                            # Calculate remaining volume needed for this product
                            total_planned = sum([info.volume1, info.volume2, info.volume3, info.volume4])
                            remaining_needed = total_planned - scheduled_volumes[product]
                            
                            if remaining_needed <= 0:
                                continue
                                
                            volume_to_schedule = min(remaining_needed, remaining_capacity)
                            
                            if add_schedule_entry(week, line, product, volume_to_schedule, info, f"Recovery Plan - Week {week}"):
                                remaining_capacity -= volume_to_schedule
                                print(f"  Scheduled {product}: {volume_to_schedule:,.0f} units")

            # Only push truly unscheduled products to Week 3
            still_unscheduled = {
                product: info for product, info in unscheduled 
                if not is_fully_scheduled(product, info)
            }
            
            if still_unscheduled:
                print("\nRemaining unscheduled products (pushing to Week 3):")
                for product, info in still_unscheduled.items():
                    remaining_volume = sum([info.volume1, info.volume2, info.volume3, info.volume4]) - scheduled_volumes[product]
                    print(f"- {product} ({info.family.value}): {remaining_volume:,.0f} units remaining")
                
                recovery_schedule = self.resolve_unscheduled_products(
                    self.schedule, 
                    {p: {
                        'planned_volume': sum([i.volume1, i.volume2, i.volume3, i.volume4]),
                        'scheduled_volume': scheduled_volumes[p]
                    } for p, i in still_unscheduled.items()}
                )
                self.schedule.extend(recovery_schedule)
        
        return self.schedule

    def group_similar_products(self, products: List[Tuple[str, ProductInfo]]) -> Dict[str, List[List[Tuple[str, ProductInfo]]]]:
        """
        Group products by priority lines, then by similarities, considering line restrictions
        """
        # First, separate products by their priority lines from Product_Summary
        line_groups = defaultdict(list)
        for product, info in products:
            # Get priority lines and their volumes
            priority_lines = [
                (info.priority1, info.volume1),
                (info.priority2, info.volume2),
                (info.priority3, info.volume3),
                (info.priority4, info.volume4)
            ]
            
            # Try assigning to each priority line in order
            assigned = False
            for priority_line, volume in priority_lines:
                if volume > 0 and self.validate_line_assignment(product, priority_line):
                    line_groups[priority_line].append((product, info))
                    assigned = True
                    break
            
            # If product couldn't be assigned to any priority line, try alternative lines
            if not assigned:
                # For light shades, try Line 2 if not already tried
                if info.shade == 'light' and 'Line 2' not in [pl[0] for pl in priority_lines]:
                    if self.validate_line_assignment(product, 'Line 2'):
                        line_groups['Line 2'].append((product, info))
                        continue
                
                # For light shades with small volume, consider Line 3 as backup
                if (info.shade == 'light' and 
                    sum([info.volume1, info.volume2, info.volume3, info.volume4]) < 5000):
                    if self.validate_line_assignment(product, 'Line 3'):
                        line_groups['Line 3'].append((product, info))
                        continue
        
        # Then group by 2-character prefix, resin, and family within each line
        final_line_groups = {}
        for line, line_products in line_groups.items():
            # First group by 2-character prefix (series)
            prefix_groups = defaultdict(list)
            for product, info in line_products:
                prefix = product[:2]  # Use first 2 characters for series
                prefix_groups[prefix].append((product, info))
            
            line_final_groups = []
            # Within each prefix group, subgroup by resin and family
            for prefix_products in prefix_groups.values():
                resin_family_groups = defaultdict(list)
                for product, info in prefix_products:
                    resin = info.resin if info.resin else 'unknown'
                    family = info.family.value if info.family else 'unknown'
                    group_key = f"{resin}_{family}"
                    resin_family_groups[group_key].append((product, info))
                
                # Add each subgroup to final groups
                for group in resin_family_groups.values():
                    if group:
                        # Sort group by shade to keep similar shades together
                        group.sort(key=lambda x: x[1].shade)
                        line_final_groups.append(group)
        
            if line_final_groups:
                final_line_groups[line] = line_final_groups
        
        return final_line_groups

    def resolve_unscheduled_products(self, schedule: List, unmet_products: Dict[str, Dict]) -> List:
        """Resolve unscheduled products with more aggressive scheduling and sequence optimization"""
        additional_schedule = []
        line_usage = defaultdict(lambda: defaultdict(float))
        
        # Track the last product on each line for each week
        last_product_on_line = defaultdict(lambda: defaultdict(dict))
        
        # Initialize line usage and last products from existing schedule
        for entry in schedule:
            week, line, product, volume = entry[0:4]
            line_usage[week][line] += volume
            last_product_on_line[week][line] = {
                'product': product,
                'family': self.product_cache[product].family,
                'resin': self.product_cache[product].resin,
                'shade': self.product_cache[product].shade
            }
        
        def get_available_capacity(week: int, line: str) -> float:
            total_capacity = LINE_CONFIG[line]['weekly_capacity']
            used_capacity = line_usage[week][line]
            return total_capacity - used_capacity
        
        # Group products by family and resin
        family_resin_groups = defaultdict(list)
        for product, data in unmet_products.items():
            info = self.product_cache[product]
            group_key = (info.family, info.resin)
            family_resin_groups[group_key].append((product, data))
        
        print("\nResolving Unscheduled Products:")
        
        # Process each family-resin group together
        for (family, resin), products in family_resin_groups.items():
            print(f"\nProcessing Family: {family}, Resin: {resin}")
            
            # Sort products by volume
            sorted_products = sorted(
                products,
                key=lambda x: x[1]['planned_volume'] - x[1]['scheduled_volume'],
                reverse=True
            )
            
            for product, data in sorted_products:
                remaining_volume = data['planned_volume'] - data['scheduled_volume']
                info = self.product_cache[product]
                print(f"\nResolving {product}: {remaining_volume:,.0f} units remaining")
                
                # Try scheduling on each line, starting with priority lines
                priority_lines = [info.priority1, info.priority2]
                all_lines = priority_lines + [line for line in LINE_CONFIG if line not in priority_lines]
                
                for line in all_lines:
                    if remaining_volume <= 0:
                        break
                    
                    week = 3  # Try week 3 first
                    
                    # Check line assignment validation
                    if not self.validate_line_assignment(product, line):
                        continue
                    
                    # Check family compatibility with last product on line
                    last_prod = last_product_on_line[week].get(line)
                    if last_prod:
                        last_family = last_prod['family']
                        last_resin = last_prod['resin']
                        
                        # Check family compatibility
                        if last_family:
                            transition_type = FAMILY_MATRIX.get(last_family, {}).get(info.family)
                            if transition_type == 'NP':
                                print(f"  Skipping {line}: Family transition not permitted")
                                continue
                        
                        # Check resin compatibility
                        if last_resin and not RESIN_MATRIX.get(last_resin, {}).get(info.resin, False):
                            print(f"  Skipping {line}: Resin transition not permitted")
                            continue
                    
                    available = get_available_capacity(week, line)
                    if available <= 0:
                        continue
                    
                    # Calculate volume to schedule
                    volume_to_schedule = min(remaining_volume, available)
                    
                    # Special handling for Line 3
                    if line == 'Line 3' and info.priority1 != 'Line 3':
                        volume_to_schedule = min(volume_to_schedule, 5000)
                    
                    if volume_to_schedule > 0:
                        # Create schedule entry
                        schedule_entry = [
                            week,
                            line,
                            product,
                            volume_to_schedule,
                            info.family,
                            info.resin,
                            info.shade,
                            f"Recovery Plan - Week {week}"
                        ]
                        
                        additional_schedule.append(schedule_entry)
                        
                        # Update tracking information
                        line_usage[week][line] += volume_to_schedule
                        last_product_on_line[week][line] = {
                            'product': product,
                            'family': info.family,
                            'resin': info.resin,
                            'shade': info.shade
                        }
                        remaining_volume -= volume_to_schedule
                        print(f"  Scheduled {volume_to_schedule:,.0f} units on {line} Week {week}")
        
        return additional_schedule

    def verify_production_targets(self, schedule: List) -> Dict[str, Dict]:
        """
        Verify if all products meet their production plan targets
        Returns: Dictionary with production status for each product
        """
        production_status = {}
        
        # Calculate total planned volume for each product
        for product, info in self.product_cache.items():
            planned_volume = sum([info.volume1, info.volume2, info.volume3, info.volume4])
            
            # Calculate actual scheduled volume
            scheduled_volume = sum(
                entry[3] for entry in schedule 
                if entry[2] == product  # entry[2] is the product code in schedule
            )
            
            # Calculate completion percentage
            completion_percentage = (scheduled_volume / planned_volume * 100) if planned_volume > 0 else 0
            
            production_status[product] = {
                'planned_volume': planned_volume,
                'scheduled_volume': scheduled_volume,
                'completion_percentage': round(completion_percentage, 2),
                'deficit': round(planned_volume - scheduled_volume, 2) if planned_volume > scheduled_volume else 0,
                'status': 'Met' if scheduled_volume >= planned_volume else 'Not Met'
            }
            
            # Log products that didn't meet their targets
            if scheduled_volume < planned_volume:
                logger.warning(
                    f"Product {product} did not meet production target: "
                    f"Planned: {planned_volume}, Achieved: {scheduled_volume}, "
                    f"Completion: {completion_percentage:.2f}%"
                )

        return production_status

    def generate_output(self, schedule: List) -> None:
        """Generate formatted output file with grouped products highlighted"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Production Schedule"
        
        current_group = None
        
        # Write schedule
        for week in range(1, 5):
            ws.append([f"Week {week}"])
            week_data = [entry for entry in schedule if entry[0] == week]
            
            for line in sorted({entry[1] for entry in week_data}):
                ws.append([line])
                line_data = [entry for entry in week_data if entry[1] == line]
                
                for entry in line_data:
                    group_id = f"{entry[5]}-{entry[2][:2]}"  # resin-prefix
                    
                    row = [
                        None, 
                        entry[2],  # product
                        entry[3],  # volume
                        f"Family: {entry[4].value if hasattr(entry[4], 'value') else entry[4]}",
                        f"Resin: {entry[5]}",
                        f"Shade: {entry[6]}",
                        entry[7]
                    ]
                    ws.append(row)
                    
                    # Add visual grouping
                    if group_id != current_group:
                        current_group = group_id
                        # Add a blank row between groups
                        ws.append([])
        
        # Write errors
        if self.errors:
            ws = wb.create_sheet("Errors")
            ws.append(["Product", "Error"])
            for product, error_list in self.errors.items():
                for error in error_list:
                    ws.append([product, error])
                    
        # Add production verification sheet
        ws = wb.create_sheet("Production Verification")
        headers = ["Product", "Planned Volume", "Scheduled Volume", 
                  "Completion %", "Deficit", "Status"]
        ws.append(headers)
        
        verification_results = self.verify_production_targets(schedule)
        for product, data in verification_results.items():
            ws.append([
                product,
                data['planned_volume'],
                data['scheduled_volume'],
                f"{data['completion_percentage']}%",
                data['deficit'],
                data['status']
            ])
        
        wb.save('formatted_final_schedule.xlsx')

    def resolve_shade_conflict(self, product: str, volume: float) -> List[Tuple[str, float]]:
        """
        Resolve shade conflicts by returning compatible lines with their efficiency ratings
        Returns: List of tuples (line_name, efficiency_rating)
        """
        product_info = self.product_cache[product]
        compatible_lines = []
        
        for line, config in LINE_CONFIG.items():
            if product_info.shade == config['shade']:
                compatible_lines.append((line, 1.0))  # 100% efficiency for matching shades
            else:
                compatible_lines.append((line, 0.7))  # 70% efficiency for mismatched shades
                
        return sorted(compatible_lines, key=lambda x: x[1], reverse=True)  # Sort by efficiency

    def optimize_final_sequence(self, schedule: List) -> List:
        """Post-process the schedule to optimize product sequences within each line and week"""
        optimized_schedule = []
        
        # Group schedule by week and line
        schedule_groups = defaultdict(lambda: defaultdict(list))
        for entry in schedule:
            week, line = entry[0], entry[1]
            schedule_groups[week][line].append(entry)
        
        def calculate_transition_score(current_product: List, next_product: List) -> float:
            """Calculate transition score between two products with strict resin compatibility"""
            # Extract resin values, assuming they're strings
            current_resin = current_product[5]
            next_resin = next_product[5]
            
            # First check: If resins are incompatible, return large negative score
            if not RESIN_MATRIX.get(current_resin, {}).get(next_resin, False):
                return float('-inf')  # Make this transition impossible
            
            score = 0.0
            
            # Family compatibility (50 points)
            # Handle both string and enum family values
            current_family = current_product[4].value if hasattr(current_product[4], 'value') else str(current_product[4])
            next_family = next_product[4].value if hasattr(next_product[4], 'value') else str(next_product[4])
            
            if current_family == next_family:
                score += 50
            elif (current_family in ['semi-glossy', 'glossy'] and 
                  next_family in ['semi-glossy', 'glossy']):
                score += 30
            
            # Resin scoring (40 points)
            if current_resin == next_resin:
                score += 40  # Same resin is best
            else:
                score += 20  # Compatible but different resin
            
            # Product series bonus (20 points)
            if current_product[2][:2] == next_product[2][:2]:
                score += 20
            
            # Debug information
            print(f"\nEvaluating transition: {current_product[2]} -> {next_product[2]}")
            print(f"  Resin transition: {current_resin} -> {next_resin}")
            print(f"  Compatible: {RESIN_MATRIX.get(current_resin, {}).get(next_resin, False)}")
            print(f"  Family transition: {current_family} -> {next_family}")
            print(f"  Score: {score}")
            
            return score
        
        def optimize_sequence(products: List) -> List:
            """Optimize sequence of products with strict resin compatibility"""
            if not products:
                return []
            
            # Group products by resin compatibility
            resin_groups = defaultdict(list)
            for product in products:
                resin = product[5].replace('Resin: ', '')
                resin_groups[resin].append(product)
            
            # Start with priority products
            priority_products = [p for p in products if "Priority Line" in p[7]]
            
            if priority_products:
                optimized = [priority_products[0]]
                # Remove used product from its resin group
                resin = priority_products[0][5].replace('Resin: ', '')
                resin_groups[resin].remove(priority_products[0])
            else:
                # Start with product from largest resin group
                start_resin = max(resin_groups.keys(), key=lambda k: len(resin_groups[k]))
                optimized = [resin_groups[start_resin].pop(0)]
            
            while any(resin_groups.values()):
                current = optimized[-1]
                current_resin = current[5].replace('Resin: ', '')
                best_score = float('-inf')
                best_next = None
                best_resin = None
                
                # Check all products in compatible resin groups
                for resin, group in resin_groups.items():
                    if not group:
                        continue
                    
                    if RESIN_MATRIX.get(current_resin, {}).get(resin, False):
                        for product in group:
                            score = calculate_transition_score(current, product)
                            if score > best_score:
                                best_score = score
                                best_next = product
                                best_resin = resin
                
                if best_next:
                    optimized.append(best_next)
                    resin_groups[best_resin].remove(best_next)
                else:
                    # If no compatible transition found, start new sequence with remaining product
                    remaining_resins = [r for r, g in resin_groups.items() if g]
                    if remaining_resins:
                        next_resin = remaining_resins[0]
                        next_product = resin_groups[next_resin].pop(0)
                        optimized.append(next_product)
                        print(f"\nWarning: No compatible transition found. Starting new sequence with {next_product[2]}")
            
            return optimized
        
        print("\nOptimizing final product sequences...")
        
        # Process each week and line
        for week in sorted(schedule_groups.keys()):
            for line in sorted(schedule_groups[week].keys()):
                products = schedule_groups[week][line]
                if len(products) > 1:  # Only optimize if there's more than one product
                    print(f"\nOptimizing Week {week} - {line}")
                    print("Original sequence:", " -> ".join([p[2] for p in products]))
                    
                    optimized_products = optimize_sequence(products)
                    
                    print("Optimized sequence:", " -> ".join([p[2] for p in optimized_products]))
                    
                    # Calculate and print transition scores
                    total_score = sum(
                        calculate_transition_score(optimized_products[i], optimized_products[i+1])
                        for i in range(len(optimized_products)-1)
                    )
                    print(f"Sequence score: {total_score:.1f}")
                    
                    optimized_schedule.extend(optimized_products)
                else:
                    optimized_schedule.extend(products)
        
        return optimized_schedule

def main():
    scheduler = ProductionScheduler()
    scheduler.load_product_data()
    scheduler.load_family_data()
    schedule = scheduler.optimize_schedule(total_weeks=2)
    
    # Verify production targets before generating output
    verification_results = scheduler.verify_production_targets(schedule)
    
    # Log overall statistics
    total_products = len(verification_results)
    met_targets = sum(1 for data in verification_results.values() if data['status'] == 'Met')
    logger.info(f"Production target achievement: {met_targets}/{total_products} products")
    
    scheduler.generate_output(schedule)

if __name__ == "__main__":
    main()